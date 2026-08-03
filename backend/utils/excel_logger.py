"""
EmberShield Excel Logger

Logs every analysis into an Excel workbook.

Each analysis (or hotspot) becomes one row.

The workbook can later be downloaded using
the /export endpoint.
"""

import os

from datetime import datetime

from openpyxl import Workbook, load_workbook

from config import EXCEL_LOG_FILE


class ExcelLogger:

    def __init__(self):

        os.makedirs("logs", exist_ok=True)

        self.file = EXCEL_LOG_FILE

        if not os.path.exists(self.file):

            workbook = Workbook()

            sheet = workbook.active

            sheet.title = "Incidents"

            sheet.append([
                "Timestamp",
                "Latitude",
                "Longitude",
                "Fire Detected",
                "Confidence",
                "Description",
                "Temperature (°C)",
                "Humidity (%)",
                "Wind Speed (km/h)",
                "Wind Direction (°)",
                "Spread Speed (km/h)",
                "Spread Direction (°)",
                "Nearest Settlement",
                "Distance (km)",
                "ETA (hrs)",
                "Severity",
                "Recommended Action",
                "Alert Message"
            ])

            workbook.save(self.file)

    def log_incident(
        self,
        latitude,
        longitude,
        sentinel,
        risk,
        commander
    ):
        """
        Append one wildfire analysis to the workbook.
        """

        workbook = load_workbook(self.file)

        sheet = workbook.active

        settlement = risk.get("nearest_settlement")

        if settlement:

            settlement_name = settlement["name"]
            distance = settlement["distance_km"]
            eta = settlement["eta_hours"]

        else:

            settlement_name = "None"
            distance = ""
            eta = ""

        sheet.append([

            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),

            latitude,

            longitude,

            sentinel["fire_detected"],

            sentinel["confidence"],

            sentinel["description"],

            risk["weather"]["temperature"],

            risk["weather"]["humidity"],

            risk["weather"]["wind_speed"],

            risk["weather"]["wind_direction"],

            risk["spread"]["speed_kmh"],

            risk["spread"]["direction"],

            settlement_name,

            distance,

            eta,

            commander["severity"],

            commander["recommended_action"],

            commander["alert_message"]

        ])

        workbook.save(self.file)

    def get_excel_path(self):
        """
        Return workbook path.
        """

        return self.file


logger = ExcelLogger()